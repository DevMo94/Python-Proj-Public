import pdfplumber
import tabulate
import openpyxl

def extract_pdf_data_after_keyword(pdf_path, keyword="Date"):
    with pdfplumber.open(pdf_path) as pdf:
        extracted_data = []
        
        # Loop through all pages
        for page in pdf.pages:
            table = page.extract_table()
            
            # Check if table is found, and search for the keyword in the first column
            if table:
                for row_index, row in enumerate(table):
                    # Start extracting data after the keyword is found
                    if keyword in row[0]:  # Assuming the keyword is in the first column
                        # Append data starting from this row onward
                        extracted_data.extend(table[row_index + 0:])
                        break  # Exit the loop after finding the keyword
            
    return extracted_data

# Example usage
pdf_path = "/Users/mohammeda/Desktop/Data_State.pdf"
keyword = "Date"  # The point where your actual data begins
extracted_data = extract_pdf_data_after_keyword(pdf_path, keyword)

tabulate_data = tabulate.tabulate(extracted_data, headers="firstrow", tablefmt="grid")
#print(tabulate_data)

def clean_extracted_data(extracted_data, output_file):
    # Load or create an Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credit"

    ws.row_dimensions[1].height = 5  # Resize first row (header)
    ws.column_dimensions['A'].width = 0.67  # Resize first column (e.g., Date column)
    
    new_headers = ["Dates", "Descriptions", "Amounts"]

    headers = extracted_data[0] if extracted_data else []

    row_num = 3
    # Write headers to the Excel sheet, starting at B2 (if needed)
    for col_num, header in enumerate(new_headers
    [0:], start=2):
        ws.cell(row=2, column=col_num, value=header)
    
    row_num = 3
    # Iterate over the data rows (skip the header row)
    for row in extracted_data[1:]:
        ws.cell(row=row_num, column=2, value = row [0])
        ws.cell(row=row_num, column=3, value = row [1])
        ws.cell(row=row_num, column=4, value = row [2])
        row_num += 1

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name (e.g., 'A', 'B', 'C')
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Save the Excel file
    wb.save(output_file)

output_excel = "/Users/mohammeda/Desktop/Credit.xlsx"

extracted_data = extract_pdf_data_after_keyword(pdf_path)
cleaned_data = clean_extracted_data(extracted_data, output_excel)
clean_extracted_data(extracted_data, output_excel)

print(f"Data extracted to excel file. Saved to {output_excel}")
